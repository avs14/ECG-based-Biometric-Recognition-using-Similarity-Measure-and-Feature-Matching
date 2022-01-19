from sewar.full_ref import mse, rmse, psnr, uqi, ssim, ergas, scc, rase, sam, msssim, vifp
from skimage.metrics import structural_similarity
import cv2 as cv
import time
import openpyxl
start= time.time()
wb = openpyxl.load_workbook('data2/outmain.xlsx')
sh1 = wb['Sheet1']


r = 1560
x = [116,119,124,212,220,217,108,100]
for i in x:
    for j in x:
        for a in [1,2,3,4,5]:
            for b in [1,2,3,4,5]:
                if a!=b:
                    img1 = cv.imread(r"data2/pics%d/%d_%d.png" %(a,i,a))
                    img2 = cv.imread(r"data2/pics%d/%d_%d.png" %(b,j,b))
                    grayA = cv.cvtColor(img1, cv.COLOR_BGR2GRAY)
                    grayB = cv.cvtColor(img2, cv.COLOR_BGR2GRAY)

                    (score, diff) = structural_similarity(grayA, grayB, full=True)


                    sh1.cell(r,1, value='%d_%d'%(i,a))
                    sh1.cell(r,2, value = '%d_%d'%(j,b))
                    sh1.cell(r, 3,value = mse(img1,img2))
                    sh1.cell(r,4, value = rmse(img1,img2))
                    sh1.cell(r,5,value=uqi(img1, img2))
                    sh1.cell(r,6,value=ergas(img1, img2))
                    sh1.cell(r, 7, value= scc(img1, img2))
                    sh1.cell(r,8,value= rase(img1, img2))
                    sh1.cell(r, 9,value=sam(img1, img2))
                    sh1.cell(r,10, value=vifp(img1, img2))
                    sh1.cell(r,11, value = score)

                    method = 'ORB'  # 'SIFT'
                    lowe_ratio = 0.89

                    if method == 'ORB':
                        finder = cv.ORB_create()
                    elif method == 'SIFT':
                        finder = cv.xfeatures2d.SIFT_create()

                    # find the keypoints and descriptors with SIFT
                    kp1, des1 = finder.detectAndCompute(img1, None)
                    kp2, des2 = finder.detectAndCompute(img2, None)

                    # BFMatcher with default params
                    bf = cv.BFMatcher()
                    matches = bf.knnMatch(des1, des2, k=2)

                    # Apply ratio test
                    good = []

                    for m, n in matches:
                        if m.distance < lowe_ratio * n.distance:
                            good.append([m])

                    sh1.cell(r, 12, value=len(good))
                    sh1.cell(r,13, value = (len(good)/500))
                    if i == j:
                        sh1.cell(r,14, value = 1)
                    else:
                        sh1.cell(r,14,value = 0)
                    r=r+1

wb.save(r'data2/outmain.xlsx')
end = time.time()
print(end-start)


#print("MSE: ", mse(blur,org))
#print("RMSE: ", rmse(blur, org))
#print("UQI: ", uqi(blur, org))
#print("ERGAS: ", ergas(blur, org))
#print("SCC: ", scc(blur, org))
#print("RASE: ", rase(blur, org))
#print("SAM: ", sam(blur, org))
#print("VIF: ", vifp(blur, org))


